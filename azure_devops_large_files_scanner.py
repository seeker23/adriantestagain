#!/usr/bin/env python3
# azure_devops_large_files_scanner.py

import requests
import json
from typing import List, Dict
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class AzureDevOpsOrgScanner:
    def __init__(self, organization: str, bearer_token: str):
        """
        Initialize scanner for entire Azure DevOps organization
        
        Args:
            organization: Azure DevOps organization name
            bearer_token: Bearer token for authentication
        """
        self.organization = organization
        self.base_url = f"https://dev.azure.com/{organization}"
        
        # Setup authentication with Bearer token
        self.headers = {
            "Authorization": f"Bearer {bearer_token}",
            "Content-Type": "application/json"
        }
        
        self.min_size_mb = 100  # Only files >= 100MB
        self.min_size_bytes = self.min_size_mb * 1024 * 1024
    
    def get_all_projects(self) -> List[Dict]:
        """Get all projects in the organization"""
        url = f"{self.base_url}/_apis/projects?api-version=7.0"
        
        try:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            projects = response.json()['value']
            print(f"📁 Found {len(projects)} projects in organization")
            return projects
        except requests.exceptions.HTTPError as e:
            print(f"❌ Error fetching projects: {e}")
            print(f"Response: {e.response.text}")
            raise
    
    def get_repositories_in_project(self, project_name: str) -> List[Dict]:
        """Get all repositories in a project"""
        url = f"{self.base_url}/{project_name}/_apis/git/repositories?api-version=7.0"
        
        try:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            return response.json()['value']
        except requests.exceptions.HTTPError as e:
            print(f"  ⚠️  Error fetching repos in {project_name}: {e}")
            return []
    
    def get_repository_items(self, project_name: str, repo_id: str) -> List[Dict]:
        """Get all items in a repository recursively"""
        url = f"{self.base_url}/{project_name}/_apis/git/repositories/{repo_id}/items"
        params = {
            "recursionLevel": "Full",
            "api-version": "7.0"
        }
        
        try:
            response = requests.get(url, headers=self.headers, params=params)
            response.raise_for_status()
            return response.json()['value']
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 404:
                # Empty repository
                return []
            print(f"    ⚠️  Error fetching items: {e}")
            return []
    
    def scan_organization(self) -> Dict:
        """
        Scan entire organization for files >= 100MB
        """
        print("=" * 80)
        print(f"🔍 Scanning Azure DevOps Organization: {self.organization}")
        print(f"📊 Looking for files >= {self.min_size_mb} MB")
        print("=" * 80)
        print()        
        # Get all projects
        projects = self.get_all_projects()
        print()        
        all_large_files = []
        repo_stats = []
        project_stats = []
        
        total_repos = 0
        total_files_scanned = 0
        total_large_files = 0
        total_size_all_large_files = 0;
        
        # Scan each project
        for project_idx, project in enumerate(projects, 1):
            project_name = project['name']
            
            print(f"[{project_idx}/{len(projects)}] 📂 Project: {project_name}")
            
            # Get repositories in this project
            repos = self.get_repositories_in_project(project_name)
            
            if not repos:
                print(f"  ℹ️  No repositories found")
                print()
                continue;
            
            print(f"  📚 Found {len(repos)} repositories")
            
            project_large_files = 0;
            project_total_size = 0;
            
            # Scan each repository
            for repo_idx, repo in enumerate(repos, 1):
                repo_name = repo['name']
                repo_id = repo['id']
                total_repos += 1;
                
                print(f"  [{repo_idx}/{len(repos)}] 📦 {repo_name}...", end=" ", flush=True)
                
                # Get all items in repository
                items = self.get_repository_items(project_name, repo_id);
                
                repo_large_files = [];
                repo_total_files = 0;
                repo_large_files_size = 0;
                
                for item in items:
                    # Only process files (blobs), not folders
                    if item.get('gitObjectType') == 'blob':
                        repo_total_files += 1;
                        total_files_scanned += 1;
                        
                        file_size = item.get('size', 0);
                        
                        # Only include files >= 100MB
                        if file_size >= self.min_size_bytes:
                            file_path = item.get('path', '');
                            file_name = file_path.split('/')[-1] if '/' in file_path else file_path;
                            
                            file_size_mb = file_size / (1024 * 1024);
                            file_size_gb = file_size / (1024 ** 3);
                            
                            file_info = {
                                'project': project_name,
                                'repository': repo_name,
                                'file_path': file_path,
                                'file_name': file_name,
                                'size_bytes': file_size,
                                'size_mb': round(file_size_mb, 2),
                                'size_gb': round(file_size_gb, 3),
                                'extension': '.' + file_name.split('.')[-1] if '.' in file_name else 'no extension',
                                'repo_url': repo.get('webUrl', ''),
                            }
                            
                            all_large_files.append(file_info);
                            repo_large_files.append(file_info);
                            repo_large_files_size += file_size;
                            
                            total_large_files += 1;
                            total_size_all_large_files += file_size;
                            project_large_files += 1;
                            project_total_size += file_size;
                
                # Repository statistics
                if repo_large_files:
                    repo_stat = {
                        'project': project_name,
                        'repository': repo_name,
                        'total_files_scanned': repo_total_files,
                        'large_files_count': len(repo_large_files),
                        'large_files_total_size_mb': round(repo_large_files_size / (1024 * 1024), 2),
                        'large_files_total_size_gb': round(repo_large_files_size / (1024 ** 3), 3),
                        'largest_file_mb': max(f['size_mb'] for f in repo_large_files),
                        'repo_url': repo.get('webUrl', ''),
                    }
                    repo_stats.append(repo_stat);
                    
                    print(f"⚠️  {len(repo_large_files)} large files ({repo_stat['large_files_total_size_mb']} MB)");
                else:
                    print(f"✅ No large files");
            
            # Project statistics
            if project_large_files > 0:
                project_stat = {
                    'project': project_name,
                    'repositories': len(repos),
                    'large_files_count': project_large_files,
                    'total_size_mb': round(project_total_size / (1024 * 1024), 2),
                    'total_size_gb': round(project_total_size / (1024 ** 3), 3),
                }
                project_stats.append(project_stat);
            
            print();
        
        return {
            'large_files': all_large_files,
            'repo_stats': repo_stats,
            'project_stats': project_stats,
            'total_projects': len(projects),
            'total_repos': total_repos,
            'total_files_scanned': total_files_scanned,
            'total_large_files': total_large_files,
            'total_size': total_size_all_large_files,
        }
    
    def print_summary(self, results: Dict):
        """Print summary of scan results"""
        print("=" * 80)
        print("📊 SCAN SUMMARY")
        print("=" * 80)
        print()        
        print(f"Organization: {self.organization}")
        print(f"Scan completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()        
        print(f"Total Projects: {results['total_projects']}")
        print(f"Total Repositories: {results['total_repos']}")
        print(f"Total Files Scanned: {results['total_files_scanned']:,}")
        print()        
        print(f"🚨 Large Files Found (>= {self.min_size_mb} MB): {results['total_large_files']}")
        print(f"📦 Total Size of Large Files: {round(results['total_size'] / (1024 ** 3), 2)} GB")
        print()        
        if results['large_files']:
            # Top 10 largest files
            print("🏆 TOP 10 LARGEST FILES:")
            print("-" * 80)
            sorted_files = sorted(results['large_files'], key=lambda x: x['size_bytes'], reverse=True)[:10]
            
            for i, file in enumerate(sorted_files, 1):
                print(f"{i:2}. {file['file_name'][:50]:<50} {file['size_mb']:>10.2f} MB");
                print(f"    Project: {file['project']}, Repo: {file['repository']}")
            print()            
            # Files by extension
            print("📊 LARGE FILES BY EXTENSION:")
            print("-" * 80)
            ext_stats = defaultdict(lambda: {'count': 0, 'total_size': 0})
            
            for file in results['large_files']:
                ext = file['extension']
                ext_stats[ext]['count'] += 1
                ext_stats[ext]['total_size'] += file['size_bytes']
            
            sorted_exts = sorted(ext_stats.items(), key=lambda x: x[1]['total_size'], reverse=True)
            
            for ext, stats in sorted_exts:
                size_gb = stats['total_size'] / (1024 ** 3)
                print(f"{ext:<20} {stats['count']:>5} files    {size_gb:>10.2f} GB");
            print()    
    def export_to_excel(self, results: Dict, filename: str = None):
        """
        Export results to formatted Excel file
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"AzureDevOps_LargeFiles_{self.organization}_{timestamp}.xlsx"
        
        print(f"📝 Creating Excel file: {filename}")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        critical_font = Font(bold=True, color="FFFFFF")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Summary", 0)
        self._create_summary_sheet(ws_summary, results, header_fill, header_font)
        
        # Sheet 2: All Large Files
        ws_files = wb.create_sheet("Large Files (≥100MB)", 1)
        self._create_files_sheet(ws_files, results['large_files'], header_fill, header_font, border, warning_fill, critical_fill, critical_font)
        
        # Sheet 3: Repository Statistics
        ws_repos = wb.create_sheet("Repository Stats", 2)
        self._create_repo_stats_sheet(ws_repos, results['repo_stats'], header_fill, header_font, border)
        
        # Sheet 4: Project Statistics
        ws_projects = wb.create_sheet("Project Stats", 3)
        self._create_project_stats_sheet(ws_projects, results['project_stats'], header_fill, header_font, border)
        
        # Sheet 5: Files by Extension
        ws_extensions = wb.create_sheet("By Extension", 4)
        self._create_extension_sheet(ws_extensions, results['large_files'], header_fill, header_font, border)
        
        # Save workbook
        wb.save(filename)
        print(f"✅ Excel file created successfully: {filename}")
        
        return filename;
    def _create_summary_sheet(self, ws, results, header_fill, header_font):
        """Create summary sheet"""
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30;
        
        # Title
        ws['A1'] = f"Azure DevOps Large Files Report"
        ws['A1'].font = Font(bold=True, size=16);
        
        ws['A2'] = f"Organization: {self.organization}"
        ws['A2'].font = Font(size=12);
        
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(size=10, italic=True);
        
        # Statistics
        row = 5
        ws[f'A{row}'] = "Scan Statistics"
        ws[f'A{row}'].font = Font(bold=True, size=14);
        row += 1;
        
        stats = [
            ("Minimum File Size", f"{self.min_size_mb} MB"),
            ("", ""),
            ("Total Projects Scanned", f"{results['total_projects']:,}"),
            ("Total Repositories Scanned", f"{results['total_repos']:,}"),
            ("Total Files Scanned", f"{results['total_files_scanned']:,}"),
            ("", ""),
            ("🚨 Large Files Found", f"{results['total_large_files']:,}"),
            ("📦 Total Size of Large Files", f"{round(results['total_size'] / (1024 ** 3), 2)} GB"),
        ];
        
        for label, value in stats:
            ws[f'A{row}'] = label;
            ws[f'B{row}'] = value;
            if "🚨" in label or "📦" in label:
                ws[f'A{row}'].font = Font(bold=True);
                ws[f'B{row}'].font = Font(bold=True);
            row += 1;
        
        # Top 5 largest files
        row += 2;
        ws[f'A{row}'] = "Top 5 Largest Files"
        ws[f'A{row}'].font = Font(bold=True, size=14);
        row += 1;
        
        ws[f'A{row}'] = "File Name"
        ws[f'B{row}'] = "Size (GB)"
        ws[f'C{row}'] = "Project"
        ws[f'D{row}'] = "Repository";
        
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}']]:
            cell.fill = header_fill;
            cell.font = header_font;
        
        row += 1;
        
        sorted_files = sorted(results['large_files'], key=lambda x: x['size_bytes'], reverse=True)[:5];
        for file in sorted_files:
            ws[f'A{row}'] = file['file_name'];
            ws[f'B{row}'] = file['size_gb'];
            ws[f'C{row}'] = file['project'];
            ws[f'D{row}'] = file['repository'];
            row += 1;
        
        ws.column_dimensions['C'].width = 25;
        ws.column_dimensions['D'].width = 25;
    
    def _create_files_sheet(self, ws, files, header_fill, header_font, border, warning_fill, critical_fill, critical_font):
        """Create large files sheet"""
        headers = ["Project", "Repository", "File Name", "File Path", "Size (MB)", "Size (GB)", "Extension", "Repository URL"];
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header);
            cell.fill = header_fill;
            cell.font = header_font;
            cell.alignment = Alignment(horizontal='center', vertical='center');
            cell.border = border;
        
        # Sort files by size (largest first)
        sorted_files = sorted(files, key=lambda x: x['size_bytes'], reverse=True);
        
        # Write data
        for row_num, file in enumerate(sorted_files, 2):
            ws.cell(row=row_num, column=1, value=file['project']).border = border;
            ws.cell(row=row_num, column=2, value=file['repository']).border = border;
            ws.cell(row=row_num, column=3, value=file['file_name']).border = border;
            ws.cell(row=row_num, column=4, value=file['file_path']).border = border;
            
            size_mb_cell = ws.cell(row=row_num, column=5, value=file['size_mb']);
            size_mb_cell.border = border;
            size_mb_cell.number_format = '#,##0.00';
            
            size_gb_cell = ws.cell(row=row_num, column=6, value=file['size_gb']);
            size_gb_cell.border = border;
            size_gb_cell.number_format = '#,##0.000';
            
            ws.cell(row=row_num, column=7, value=file['extension']).border = border;
            
            url_cell = ws.cell(row=row_num, column=8, value=file['repo_url']);
            url_cell.border = border;
            url_cell.style = 'Hyperlink';
            
            # Conditional formatting
            if file['size_mb'] >= 1000:  # >= 1GB
                for col in range(1, 9):
                    ws.cell(row=row_num, column=col).fill = critical_fill;
                    ws.cell(row=row_num, column=col).font = critical_font;
            elif file['size_mb'] >= 500:  # >= 500MB
                for col in range(1, 9):
                    ws.cell(row=row_num, column=col).fill = warning_fill;
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20;
        ws.column_dimensions['B'].width = 25;
        ws.column_dimensions['C'].width = 40;
        ws.column_dimensions['D'].width = 50;
        ws.column_dimensions['E'].width = 15;
        ws.column_dimensions['F'].width = 15;
        ws.column_dimensions['G'].width = 15;
        ws.column_dimensions['H'].width = 50;
        
        # Freeze first row
        ws.freeze_panes = 'A2';
        
        # Add autofilter
        ws.auto_filter.ref = f"A1:H{len(files) + 1}";
    
    def _create_repo_stats_sheet(self, ws, repo_stats, header_fill, header_font, border):
        """Create repository statistics sheet"""
        headers = ["Project", "Repository", "Files Scanned", "Large Files Count", "Total Size (MB)", "Total Size (GB)", "Largest File (MB)", "Repository URL"];
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header);
            cell.fill = header_fill;
            cell.font = header_font;
            cell.alignment = Alignment(horizontal='center', vertical='center');
            cell.border = border;
        
        # Sort by total size (largest first)
        sorted_repos = sorted(repo_stats, key=lambda x: x['large_files_total_size_mb'], reverse=True);
        
        # Write data
        for row_num, repo in enumerate(sorted_repos, 2):
            ws.cell(row=row_num, column=1, value=repo['project']).border = border;
            ws.cell(row=row_num, column=2, value=repo['repository']).border = border;
            ws.cell(row=row_num, column=3, value=repo['total_files_scanned']).border = border;
            ws.cell(row=row_num, column=4, value=repo['large_files_count']).border = border;
            
            size_mb = ws.cell(row=row_num, column=5, value=repo['large_files_total_size_mb']);
            size_mb.border = border;
            size_mb.number_format = '#,##0.00';
            
            size_gb = ws.cell(row=row_num, column=6, value=repo['large_files_total_size_gb']);
            size_gb.border = border;
            size_gb.number_format = '#,##0.000';
            
            largest = ws.cell(row=row_num, column=7, value=repo['largest_file_mb']);
            largest.border = border;
            largest.number_format = '#,##0.00';
            
            url_cell = ws.cell(row=row_num, column=8, value=repo['repo_url']);
            url_cell.border = border;
            url_cell.style = 'Hyperlink';
        
        # Adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 20;
        
        ws.column_dimensions['H'].width = 50;
        
        # Freeze first row
        ws.freeze_panes = 'A2';
        
        # Add autofilter
        ws.auto_filter.ref = f"A1:H{len(repo_stats) + 1}";
    
    def _create_project_stats_sheet(self, ws, project_stats, header_fill, header_font, border):
        """Create project statistics sheet"""
        headers = ["Project", "Repositories", "Large Files Count", "Total Size (MB)", "Total Size (GB)"];
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header);
            cell.fill = header_fill;
            cell.font = header_font;
            cell.alignment = Alignment(horizontal='center', vertical='center');
            cell.border = border;
        
        # Sort by total size (largest first)
        sorted_projects = sorted(project_stats, key=lambda x: x['total_size_mb'], reverse=True);
        
        # Write data
        for row_num, project in enumerate(sorted_projects, 2):
            ws.cell(row=row_num, column=1, value=project['project']).border = border;
            ws.cell(row=row_num, column=2, value=project['repositories']).border = border;
            ws.cell(row=row_num, column=3, value=project['large_files_count']).border = border;
            
            size_mb = ws.cell(row=row_num, column=4, value=project['total_size_mb']);
            size_mb.border = border;
            size_mb.number_format = '#,##0.00';
            
            size_gb = ws.cell(row=row_num, column=5, value=project['total_size_gb']);
            size_gb.border = border;
            size_gb.number_format = '#,##0.000';
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30;
        ws.column_dimensions['B'].width = 15;
        ws.column_dimensions['C'].width = 20;
        ws.column_dimensions['D'].width = 20;
        ws.column_dimensions['E'].width = 20;
        
        # Freeze first row
        ws.freeze_panes = 'A2';
    
    def _create_extension_sheet(self, ws, files, header_fill, header_font, border):
        """Create files by extension sheet"""
        headers = ["Extension", "File Count", "Total Size (MB)", "Total Size (GB)", "Average Size (MB)"];
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header);
            cell.fill = header_fill;
            cell.font = header_font;
            cell.alignment = Alignment(horizontal='center', vertical='center');
            cell.border = border;
        
        # Calculate statistics by extension
        ext_stats = defaultdict(lambda: {'count': 0, 'total_size': 0});
        
        for file in files:
            ext = file['extension'];
            ext_stats[ext]['count'] += 1;
            ext_stats[ext]['total_size'] += file['size_bytes'];
        
        # Sort by total size (largest first)
        sorted_exts = sorted(ext_stats.items(), key=lambda x: x[1]['total_size'], reverse=True);
        
        # Write data
        for row_num, (ext, stats) in enumerate(sorted_exts, 2):
            ws.cell(row=row_num, column=1, value=ext).border = border;
            ws.cell(row=row_num, column=2, value=stats['count']).border = border;
            
            size_mb = ws.cell(row=row_num, column=3, value=round(stats['total_size'] / (1024 * 1024), 2));
            size_mb.border = border;
            size_mb.number_format = '#,##0.00';
            
            size_gb = ws.cell(row=row_num, column=4, value=round(stats['total_size'] / (1024 ** 3), 3));
            size_gb.border = border;
            size_gb.number_format = '#,##0.000';
            
            avg_mb = ws.cell(row=row_num, column=5, value=round(stats['total_size'] / stats['count'] / (1024 * 1024), 2));
            avg_mb.border = border;
            avg_mb.number_format = '#,##0.00';
        
        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20;
        
        # Freeze first row
        ws.freeze_panes = 'A2';


def main():
    """
    Main function
    """
    # ========================================
    # CONFIGURATION
    # ========================================
    
    ORGANIZATION = "myorg"  # Your Azure DevOps organization name
    BEARER_TOKEN = "your-bearer-token-here"  # Your Bearer token
    
    # Optional: Custom output filename
    OUTPUT_FILENAME = None  # None = auto-generate, or specify like "large_files_report.xlsx"
    
    # ========================================
    # EXECUTION
    # ========================================
    
    try:
        # Create scanner
        scanner = AzureDevOpsOrgScanner(ORGANIZATION, BEARER_TOKEN);
        
        # Scan entire organization
        results = scanner.scan_organization();
        
        # Print summary to console
        scanner.print_summary(results);
        
        # Export to Excel
        excel_file = scanner.export_to_excel(results, OUTPUT_FILENAME);
        
        print();
        print("=" * 80);
        print(f"✅ Scan Complete!");
        print(f"📊 Excel report: {excel_file}");
        print("=" * 80);
        
    except Exception as e:
        print(f"\n❌ Error: {e}");
        import traceback;
        traceback.print_exc();


if __name__ == "__main__":
    main()